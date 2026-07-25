"""
Product bulk import service layer.

Provides Excel parsing, data validation, and batch write functionality.
"""
from typing import Any
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.product import Product, ProductCategory, ProductUnit
from app.services.product_service import check_barcode_exists

# Excel column name -> Product field mapping
COLUMN_MAP: dict[str, str] = {
    "商品名称": "name",
    "品牌": "brand",
    "分类": "category",
    "规格": "spec",
    "单位": "unit",
    "条码": "barcode",
    "SKU编码": "sku_code",
    "成本价": "cost_price",
    "零售价": "retail_price",
    "批发价": "wholesale_price",
    "安全库存": "safety_stock",
    "备注": "remark",
}

# Field name -> Chinese label (for error messages)
FIELD_CN: dict[str, str] = {v: k for k, v in COLUMN_MAP.items()}

VALID_CATEGORIES: set[str] = {e.value for e in ProductCategory}
VALID_UNITS: set[str] = {e.value for e in ProductUnit}

# Template header order
TEMPLATE_HEADERS: list[str] = [
    "商品名称", "品牌", "分类", "规格", "单位",
    "条码", "SKU编码", "成本价", "零售价", "批发价",
    "安全库存", "备注",
]


def generate_template() -> bytes:
    """Generate product import template Excel (headers + example rows)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "商品导入模板"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # Example row style
    example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Write headers
    for col_idx, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write example data
    examples = [
        ["示例-精华液", "示例品牌", "护肤", "30ml", "瓶", "", "", 80, 299, 260, 20, "示例数据"],
        ["示例-口红", "示例品牌", "彩妆", "3.5g", "支", "", "", 60, 199, 170, 15, ""],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = example_fill

    # Column widths
    col_widths = [16, 12, 10, 10, 8, 16, 14, 10, 10, 10, 10, 16]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_excel(file_content: bytes) -> list[dict[str, Any]]:
    """Parse Excel file, return list of valid row dicts."""
    wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, values_only=True)

    # Read header row
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return []

    headers = [str(h).strip() if h else "" for h in header_row]

    # Build column index mapping: column index -> field name
    col_index: dict[int, str] = {}
    for excel_name, field_name in COLUMN_MAP.items():
        if excel_name in headers:
            col_index[headers.index(excel_name)] = field_name

    if not col_index:
        wb.close()
        return []

    rows: list[dict[str, Any]] = []
    for row in rows_iter:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        data: dict[str, Any] = {}
        for idx, field_name in col_index.items():
            val = row[idx] if idx < len(row) else None
            data[field_name] = val
        rows.append(data)

    wb.close()
    return rows


def validate_row(row: dict[str, Any], row_num: int) -> list[str]:
    """校验单行数据，返回错误信息列表"""
    errors: list[str] = []

    # Product name is required
    name = _str_val(row.get("name"))
    if not name:
        errors.append(f"第{row_num}行：商品名称不能为空")
    elif len(name) > 200:
        errors.append(f"第{row_num}行：商品名称不能超过200字符")

    # Category validation
    category = _str_val(row.get("category"))
    if category and category not in VALID_CATEGORIES:
        errors.append(
            f"第{row_num}行：分类\"{category}\"无效，可选值：{'、'.join(sorted(VALID_CATEGORIES))}"
        )

    # Unit validation
    unit = _str_val(row.get("unit"))
    if unit and unit not in VALID_UNITS:
        errors.append(
            f"第{row_num}行：单位\"{unit}\"无效，可选值：{'、'.join(sorted(VALID_UNITS))}"
        )

    # Price validation
    for price_field in ("cost_price", "retail_price", "wholesale_price"):
        val = row.get(price_field)
        if val is not None and val != "" and str(val).strip() != "":
            try:
                p = float(val)
                if p < 0:
                    errors.append(
                        f"第{row_num}行：{FIELD_CN.get(price_field, price_field)}不能为负数"
                    )
            except (ValueError, TypeError):
                errors.append(
                    f"第{row_num}行：{FIELD_CN.get(price_field, price_field)}格式不正确"
                )

    # Safety stock validation
    val = row.get("safety_stock")
    if val is not None and val != "" and str(val).strip() != "":
        try:
            s = int(float(val))
            if s < 0:
                errors.append(f"第{row_num}行：安全库存不能为负数")
        except (ValueError, TypeError):
            errors.append(f"第{row_num}行：安全库存格式不正确")

    # Barcode length validation
    barcode = _str_val(row.get("barcode"))
    if barcode and len(barcode) > 50:
        errors.append(f"第{row_num}行：条码不能超过50字符")

    # SKU code length validation
    sku = _str_val(row.get("sku_code"))
    if sku and len(sku) > 20:
        errors.append(f"第{row_num}行：SKU编码不能超过20字符")

    # Spec length validation
    spec = _str_val(row.get("spec"))
    if spec and len(spec) > 50:
        errors.append(f"第{row_num}行：规格不能超过50字符")

    return errors


def preview_import(rows: list[dict[str, Any]]) -> list[dict]:
    """预览导入数据，返回带校验结果的行列表"""
    results: list[dict] = []
    for i, row in enumerate(rows):
        row_num = i + 2  # Row 1 is header, data starts from row 2
        errors = validate_row(row, row_num)

        # Build display data
        display: dict[str, Any] = {}
        for field_name in COLUMN_MAP.values():
            val = row.get(field_name)
            if val is None:
                display[field_name] = ""
            elif isinstance(val, (int, float)):
                display[field_name] = val
            else:
                display[field_name] = str(val).strip()

        results.append({
            "row_number": row_num,
            "data": display,
            "errors": errors,
            "valid": len(errors) == 0,
        })

    return results


def batch_import(db: Session, rows: list[dict[str, Any]]) -> dict:
    """逐行批量导入商品，单行失败不影响其他行"""
    from app.services.product_service import _generate_sku_code

    total = len(rows)
    success = 0
    failed: list[dict] = []

    for i, row in enumerate(rows):
        row_num = i + 2

        # Validate
        errors = validate_row(row, row_num)
        if errors:
            failed.append({"row_number": row_num, "errors": errors})
            continue

        # Skip completely empty rows
        name = _str_val(row.get("name"))
        if not name:
            continue

        try:
            brand = _str_val(row.get("brand"))
            category = _str_val(row.get("category")) or "其他"
            unit = _str_val(row.get("unit")) or "个"
            barcode = _str_val(row.get("barcode")) or None
            sku_code = _str_val(row.get("sku_code")) or None
            spec = _str_val(row.get("spec"))
            remark = _str_val(row.get("remark"))

            cost_price = _float_val(row.get("cost_price"), 0)
            retail_price = _float_val(row.get("retail_price"), 0)
            wholesale_price = _float_val(row.get("wholesale_price"), 0)
            safety_stock = _int_val(row.get("safety_stock"), 10)

            # Convert empty barcode to None
            if barcode and barcode.strip() == "":
                barcode = None

            # Barcode uniqueness check
            if barcode and check_barcode_exists(db, barcode):
                failed.append({
                    "row_number": row_num,
                    "errors": [f"第{row_num}行：条码\"{barcode}\"已存在"],
                })
                continue

            # Auto-generate SKU code
            if not sku_code:
                sku_code = _generate_sku_code(db, brand)

            product = Product(
                name=name,
                brand=brand,
                category=category,
                unit=unit,
                barcode=barcode,
                sku_code=sku_code,
                spec=spec or "",
                cost_price=cost_price,
                retail_price=retail_price,
                wholesale_price=wholesale_price,
                safety_stock=safety_stock,
                remark=remark or "",
                status="在售",
            )
            db.add(product)
            db.flush()

            # Auto-generate barcode (when not provided)
            if not barcode:
                product.barcode = f"BH{product.id:08d}"

            db.commit()
            success += 1

        except Exception as e:
            db.rollback()
            failed.append({
                "row_number": row_num,
                "errors": [f"第{row_num}行：导入失败 - {str(e)}"],
            })

    return {
        "total": total,
        "success": success,
        "failed": failed,
    }


def _str_val(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _float_val(val, default: float = 0.0) -> float:
    if val is None or str(val).strip() == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _int_val(val, default: int = 0) -> int:
    if val is None or str(val).strip() == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default
