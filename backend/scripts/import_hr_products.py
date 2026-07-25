"""
赫莲娜产品数据导入脚本
Data source: hr-scraper/output/products.json + products(梳理后)).xlsx
Target: product table
"""

import json
import sys
import os
import sqlite3
from pathlib import Path

DB_PATH = str(Path(__file__).resolve().parent.parent / "data" / "business.db")
JSON_PATH = r"D:\资源\赫莲娜产品信息\hr-scraper\output\products.json"
XLSX_PATH = r"D:\资源\赫莲娜产品信息\hr-scraper\output\products（梳理后））.xlsx"

try:
    import openpyxl
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    sys.exit(1)


def infer_unit(spec, name):
    combined = f"{name} {spec}".lower()
    if any(k in combined for k in ["面膜", "套组", "套装"]):
        return "盒"
    if any(k in combined for k in ["眼霜", "精华乳", "洁面乳", "精华露", "精萃露", "精华液"]):
        return "支"
    if spec and (spec.endswith("ml") or spec.endswith("g")):
        return "瓶"
    if any(k in combined for k in ["美容液"]):
        return "瓶"
    return "瓶"


def load_excel_prices(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    id_idx = headers.index("ID")
    price_idx = headers.index("价格")

    prices = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid, price = row[id_idx], row[price_idx]
        if pid and price:
            try:
                prices[pid] = float(price)
            except (ValueError, TypeError):
                pass
    wb.close()
    return prices


def load_sku_data(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    skus = []
    for product in data["products"]:
        for sku in product["basicSkuList"]:
            specs = sku.get("specifications", [])
            spec_value = specs[0]["value"] if specs else ""
            images = sku.get("images", [])
            skus.append({
                "product_id": product["id"],
                "name": sku["goodsName"],
                "barcode": sku["code"].strip(),
                "spec": spec_value,
                "main_image": sku.get("mainImage", "") or (images[0] if images else ""),
            })
    return skus


def import_products():
    prices = load_excel_prices(XLSX_PATH)
    skus = load_sku_data(JSON_PATH)

    print(f"Excel 价格条目: {len(prices)}")
    print(f"JSON SKU 条目: {len(skus)}")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    # Ensure product table exists
    conn.execute("""
        CREATE TABLE IF NOT EXISTS product (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name VARCHAR(200) NOT NULL,
            barcode VARCHAR(50) UNIQUE,
            spec VARCHAR(100),
            category VARCHAR(20) NOT NULL DEFAULT '其他',
            unit VARCHAR(10) NOT NULL DEFAULT '个',
            image_url VARCHAR(500),
            cost_price FLOAT NOT NULL DEFAULT 0,
            retail_price FLOAT NOT NULL DEFAULT 0,
            wholesale_price FLOAT DEFAULT 0,
            safety_stock INTEGER DEFAULT 10,
            status VARCHAR(10) NOT NULL DEFAULT '在售',
            remark TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    created = 0
    skipped = 0
    no_price = 0

    for sku in skus:
        barcode = sku["barcode"]
        if not barcode:
            skipped += 1
            continue

        # Check duplicate
        cur = conn.execute("SELECT id FROM product WHERE barcode = ?", (barcode,))
        if cur.fetchone():
            skipped += 1
            continue

        price = prices.get(sku["product_id"], 0)
        if price == 0:
            no_price += 1

        unit = infer_unit(sku["spec"], sku["name"])

        conn.execute(
            """INSERT INTO product (name, barcode, spec, category, unit, image_url,
               retail_price, cost_price, wholesale_price, safety_stock, status, remark)
               VALUES (?, ?, ?, '护肤', ?, ?,  ?, 0, 0, 10, '在售', '赫莲娜')""",
            (sku["name"], barcode, sku["spec"], unit, sku["main_image"], price),
        )
        created += 1

    conn.commit()
    cur = conn.execute("SELECT COUNT(*) FROM product")
    total = cur.fetchone()[0]
    conn.close()

    print(f"\n导入完成:")
    print(f"  新增: {created}")
    print(f"  跳过 (重复/无条码): {skipped}")
    print(f"  缺价格: {no_price}")
    print(f"  数据库 product 表总数: {total}")


if __name__ == "__main__":
    import_products()
